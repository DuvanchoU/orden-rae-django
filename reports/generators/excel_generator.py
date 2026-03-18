from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

class ExcelReportGenerator:
    """Generador base para reportes en Excel"""
    
    HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="FF2DD4BF", end_color="FF2DD4BF", fill_type="solid")
    CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def __init__(self, filename, headers, data, sheet_name="Reporte"):
        self.filename = filename
        self.headers = headers
        self.data = data
        self.sheet_name = sheet_name
        self.wb = Workbook()
    
    def _apply_header_style(self, cell):
        cell.font = self.HEADER_FONT
        cell.fill = self.HEADER_FILL
        cell.alignment = self.CELL_ALIGNMENT
        cell.border = self.THIN_BORDER
    
    def _apply_cell_style(self, cell):
        cell.alignment = self.CELL_ALIGNMENT
        cell.border = self.THIN_BORDER
    
    def generate(self):
        ws = self.wb.active
        ws.title = self.sheet_name
        
        # Encabezados
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header.upper())
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col)].width = len(str(header)) + 2
        
        # Datos
        for row_idx, row_data in enumerate(self.data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                self._apply_cell_style(cell)
        
        ws.freeze_panes = 'A2'
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}.xlsx"'
        self.wb.save(response)
        return response